"""
excel_reader.py

Reads a SonarQube-exported Excel workbook with the following layout:

    Sheet 1 — Instructions   (ignored by this module)
    Sheet 2 — Rules          Master list of Sonar rules; one row per rule.
    Sheet 3…N — <RuleKey>   One sheet per rule; every row is an individual issue.

Expected column headers
-----------------------
Rules sheet (Sheet 2):
    key        – Sonar rule key  (e.g. cs-S1006)
    sheetIden  – Name of the per-rule sheet that holds the issues for this rule
    name       – Human-readable rule name
    severity   – BLOCKER | CRITICAL | MAJOR | MINOR | INFO
    language   – Programming language (cs, java, py, …)
    action     – Typically "noaction" until the platform processes the row
    count      – Number of open issues for this rule

Per-rule sheets (Sheet 3…N):
    key        – Sonar issue key (UUID)
    severity   – BLOCKER | CRITICAL | MAJOR | MINOR | INFO
    message    – Issue-specific description / fix hint
    line       – Source line number
    component  – <RepoKey>:<Branch>:<relativeFilePath>  OR  <ProjectKey>:<relativeFilePath>
    assigneeU  – Sonar assignee username (optional)
    assignee   – Sonar assignee display name (optional)
    status     – OPEN | CONFIRMED | RESOLVED | CLOSED

Usage
-----
    from ingestion.excel_reader import ExcelReader, RuleInfo, IssueModel

    reader = ExcelReader("data/issues.xlsx")
    rules  = reader.load_rules()        # list[RuleInfo]
    issues = reader.load_all_issues()   # list[IssueModel]  (status == OPEN only)
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class RuleInfo:
    """Represents one row from the Rules sheet (Sheet 2)."""
    key: str          # e.g. "cs-S1006"
    sheet_iden: str   # name of the per-rule worksheet
    name: str         # human-readable description
    severity: str     # BLOCKER | CRITICAL | MAJOR | MINOR | INFO
    language: str     # cs | java | py | …
    action: str       # typically "noaction"
    count: int        # number of open issues


@dataclass
class IssueModel:
    """Represents one row from a per-rule sheet (Sheet 3…N)."""
    key: str                       # Sonar issue UUID
    severity: str                  # BLOCKER | CRITICAL | MAJOR | MINOR | INFO
    message: str                   # Description of the problem
    line: int                      # Source line number
    component: str                 # Raw component field (parsed by ComponentParser)
    assignee_username: str         # Sonar assignee login (may be empty)
    assignee_display: str          # Sonar assignee display name (may be empty)
    status: str                    # OPEN | CONFIRMED | RESOLVED | CLOSED
    rule: Optional[RuleInfo] = field(default=None, repr=False)  # parent rule


# Severity order (higher index = higher severity)
_SEVERITY_ORDER = {
    "INFO": 0,
    "MINOR": 1,
    "MAJOR": 2,
    "CRITICAL": 3,
    "BLOCKER": 4,
}


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _cell_str(value) -> str:
    """Return a stripped string regardless of the cell value type."""
    if value is None:
        return ""
    return str(value).strip()


def _cell_int(value, default: int = 0) -> int:
    """Return an integer from a cell value, falling back to *default*."""
    if value is None:
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _normalise_headers(row) -> dict[str, int]:
    """
    Map each header string (lower-cased, stripped) to its 0-based column index.

    Parameters
    ----------
    row : tuple
        The header row from an openpyxl sheet (``ws[1]``).
    """
    return {
        _cell_str(cell.value).lower(): idx
        for idx, cell in enumerate(row)
        if cell.value is not None
    }


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------

class ExcelReader:
    """
    Reads a SonarQube issue export workbook and materialises structured objects.

    Parameters
    ----------
    path : str | Path
        Path to the ``.xlsx`` workbook.
    open_statuses : set[str] | None
        Issue status values considered "actionable".
        Defaults to ``{"OPEN", "CONFIRMED"}``.
    severity_threshold : str | None
        Minimum severity to include.  Issues below this threshold are skipped.
        Accepted values: INFO, MINOR, MAJOR, CRITICAL, BLOCKER.
        Defaults to ``None`` (no filter).
    """

    _RULES_SHEET_INDEX = 1  # 0-based; Sheet 2 in Excel (Sheet 1 = instructions)
    _ISSUES_SHEET_START = 2  # 0-based; Sheet 3 onwards

    def __init__(
        self,
        path: str | Path,
        open_statuses: set[str] | None = None,
        severity_threshold: str | None = None,
    ) -> None:
        self._path = Path(path)
        self._open_statuses: set[str] = open_statuses or {"OPEN", "CONFIRMED", "ACCEPTED"}
        self._severity_min: int = (
            _SEVERITY_ORDER.get(severity_threshold.upper(), 0)
            if severity_threshold
            else 0
        )
        self._wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def load_rules(self) -> list[RuleInfo]:
        """
        Parse the Rules sheet (Sheet 2) and return a list of :class:`RuleInfo`.

        Rows where ``key`` is empty are silently skipped.
        """
        sheets = self._wb.sheetnames
        if len(sheets) < 2:
            raise ValueError(
                f"Workbook '{self._path}' has fewer than 2 sheets; "
                "expected Sheet 1 = Instructions, Sheet 2 = Rules."
            )

        ws = self._wb[sheets[self._RULES_SHEET_INDEX]]
        rows = list(ws.iter_rows())
        if not rows:
            logger.warning("Rules sheet is empty.")
            return []

        headers = _normalise_headers(rows[0])
        _require_headers(headers, {"key", "severity", "language"}, sheet_name="Rules")

        rules: list[RuleInfo] = []
        for row in rows[1:]:
            cells = [_cell_str(c.value) for c in row]
            key = _get(cells, headers, "key")
            if not key:
                continue  # blank row

            rules.append(
                RuleInfo(
                    key=key,
                    sheet_iden=(
                        _get(cells, headers, "sheetidentifier")
                        or _get(cells, headers, "sheetiden")
                        or key
                    ),
                    name=_get(cells, headers, "name"),
                    severity=_get(cells, headers, "severity").upper(),
                    language=_get(cells, headers, "language").lower(),
                    action=_get(cells, headers, "action"),
                    count=_cell_int(_raw(row, headers, "count")),
                )
            )

        logger.info("Loaded %d rules from '%s'.", len(rules), self._path.name)
        return rules

    def load_issues_for_rule(self, rule: RuleInfo) -> list[IssueModel]:
        """
        Parse the per-rule sheet identified by *rule.sheet_iden* and return a
        list of :class:`IssueModel` whose status is in ``open_statuses`` and
        whose severity meets the configured threshold.

        The method tries the following sheet names in order and skips gracefully
        if none are found or the sheet cannot be read:

        1. ``rule.sheet_iden``
        2. ``rule.key``

        Parameters
        ----------
        rule : RuleInfo
            The rule whose sheet should be parsed.
        """
        # Resolve which sheet to read, trying sheet_iden then key as fallback
        sheet_name = self._resolve_sheet_name(rule)
        if sheet_name is None:
            return []

        try:
            return self._read_issue_sheet(sheet_name, rule)
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "Rule '%s': could not read sheet '%s' — %s. Skipping.",
                rule.key, sheet_name, exc,
            )
            return []

    def _resolve_sheet_name(self, rule: RuleInfo) -> str | None:
        """Return the first existing sheet name for *rule*, or ``None``."""
        available = set(self._wb.sheetnames)
        for candidate in (rule.sheet_iden, rule.key):
            if candidate and candidate in available:
                logger.debug("Rule '%s': using sheet '%s'.", rule.key, candidate)
                return candidate
        logger.warning(
            "Rule '%s': no matching sheet found (tried '%s', '%s'). Skipping.",
            rule.key, rule.sheet_iden, rule.key,
        )
        return None

    def _read_issue_sheet(self, sheet_name: str, rule: RuleInfo) -> list[IssueModel]:
        """Read *sheet_name* and return filtered :class:`IssueModel` objects."""
        ws = self._wb[sheet_name]
        rows = list(ws.iter_rows())
        if not rows:
            logger.warning("Sheet '%s' is empty.", sheet_name)
            return []

        headers = _normalise_headers(rows[0])

        # Validate required columns; skip sheet if any are absent
        missing = {"key", "severity", "status"} - headers.keys()
        if missing:
            logger.warning(
                "Sheet '%s' is missing column(s) %s — skipping.",
                sheet_name, sorted(missing),
            )
            return []

        issues: list[IssueModel] = []
        for row in rows[1:]:
            cells = [_cell_str(c.value) for c in row]
            key = _get(cells, headers, "key")
            if not key:
                continue

            status = _get(cells, headers, "status").upper()
            if status not in self._open_statuses:
                continue

            severity = _get(cells, headers, "severity").upper()
            if _SEVERITY_ORDER.get(severity, 0) < self._severity_min:
                continue

            issues.append(
                IssueModel(
                    key=key,
                    severity=severity,
                    message=_get(cells, headers, "message"),
                    line=_cell_int(_raw(row, headers, "line"), default=1),
                    component=_get(cells, headers, "component"),
                    assignee_username=(
                        _get(cells, headers, "assigneeusername")
                        or _get(cells, headers, "assigneeu")
                    ),
                    assignee_display=_get(cells, headers, "assignee"),
                    status=status,
                    rule=rule,
                )
            )

        logger.info(
            "Rule '%s': %d actionable issue(s) in sheet '%s'.",
            rule.key, len(issues), sheet_name,
        )
        return issues

    def load_all_issues(
        self,
        allowed_rules: set[str] | None = None,
    ) -> list[IssueModel]:
        """
        Load all actionable issues across every rule sheet.

        Parameters
        ----------
        allowed_rules : set[str] | None
            If supplied, only issues belonging to rules in this set are returned.
            Accepts rule keys (e.g. ``{"cs-S1006", "cs-S1110"}``).
            Pass ``None`` to load all rules.

        Returns
        -------
        list[IssueModel]
            Issues ordered by severity (highest first), then by rule key, then
            by file / line.
        """
        rules = self.load_rules()
        all_issues: list[IssueModel] = []

        for rule in rules:
            if allowed_rules and rule.key not in allowed_rules:
                logger.debug("Rule '%s' not in allow-list; skipping.", rule.key)
                continue
            try:
                all_issues.extend(self.load_issues_for_rule(rule))
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "Rule '%s': unexpected error while loading issues — %s. Skipping.",
                    rule.key, exc,
                )

        # Sort: highest severity first, then rule key, then line
        all_issues.sort(
            key=lambda i: (
                -_SEVERITY_ORDER.get(i.severity, 0),
                i.rule.key if i.rule else "",
                i.line,
            )
        )

        logger.info(
            "Total actionable issues loaded: %d (across %d rule(s)).",
            len(all_issues), len(rules),
        )
        return all_issues

    def close(self) -> None:
        """Close the underlying workbook (frees file handle)."""
        self._wb.close()

    def __enter__(self) -> "ExcelReader":
        return self

    def __exit__(self, *_) -> None:
        self.close()


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _get(cells: list[str], headers: dict[str, int], col: str) -> str:
    """Return the string value in *cells* for column *col*, or empty string."""
    idx = headers.get(col.lower())
    if idx is None or idx >= len(cells):
        return ""
    return cells[idx]


def _raw(row, headers: dict[str, int], col: str):
    """Return the raw cell value for *col* in *row*."""
    idx = headers.get(col.lower())
    if idx is None or idx >= len(row):
        return None
    return row[idx].value


def _require_headers(
    headers: dict[str, int],
    required: set[str],
    sheet_name: str,
) -> None:
    """Raise *ValueError* if any of the *required* columns are absent."""
    missing = required - headers.keys()
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' is missing expected column(s): "
            f"{sorted(missing)}. Found: {sorted(headers.keys())}"
        )
