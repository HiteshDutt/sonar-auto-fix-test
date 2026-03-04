"""
Tests for ingestion/excel_reader.py

Uses an in-memory openpyxl workbook to avoid needing real .xlsx files on disk.
"""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from ingestion.excel_reader import ExcelReader, RuleInfo, IssueModel


# ---------------------------------------------------------------------------
# Fixtures — in-memory workbooks
# ---------------------------------------------------------------------------

def _make_workbook(
    rules: list[tuple],
    issues_per_rule: dict[str, list[tuple]],
) -> BytesIO:
    """
    Build an in-memory .xlsx file matching the expected sheet layout.

    Parameters
    ----------
    rules : list[tuple]
        Rows for the Rules sheet (Sheet 2), excluding the header.
        Each tuple: (key, sheetIden, name, severity, language, action, count)
    issues_per_rule : dict[str, list[tuple]]
        Mapping from sheet name → list of issue rows.
        Each tuple: (key, severity, message, line, component, assigneeU, assignee, status)
    """
    wb = openpyxl.Workbook()

    # Sheet 1 — Instructions (ignored)
    inst = wb.active
    inst.title = "Instructions"
    inst.append(["This sheet is ignored by the auto-fix tool."])

    # Sheet 2 — Rules
    rules_ws = wb.create_sheet("Rules")
    rules_ws.append(["key", "sheetIden", "name", "severity", "language", "action", "count"])
    for row in rules:
        rules_ws.append(list(row))

    # Sheet 3…N — per-rule issues
    for sheet_name, issue_rows in issues_per_rule.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(["key", "severity", "message", "line", "component", "assigneeU", "assignee", "status"])
        for row in issue_rows:
            ws.append(list(row))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture()
def simple_workbook(tmp_path: Path) -> Path:
    """A minimal workbook with 2 rules and 3 issues total."""
    buf = _make_workbook(
        rules=[
            ("cs-S1006", "S1006", "Method overrides should call super()", "CRITICAL", "cs", "noaction", 1),
            ("cs-S1110", "S1110", "Redundant type casts", "MAJOR", "cs", "noaction", 2),
        ],
        issues_per_rule={
            "S1006": [
                ("uuid-001", "CRITICAL", "Add a call to super()", 42, "myrepo:main:src/Foo.cs", "user1", "Alice", "OPEN"),
            ],
            "S1110": [
                ("uuid-002", "MAJOR", "Remove redundant cast", 10, "myrepo:main:src/Bar.cs", "", "", "OPEN"),
                ("uuid-003", "MAJOR", "Remove redundant cast", 20, "myrepo:main:src/Baz.cs", "", "", "RESOLVED"),
            ],
        },
    )
    p = tmp_path / "issues.xlsx"
    p.write_bytes(buf.read())
    return p


# ---------------------------------------------------------------------------
# Tests — load_rules
# ---------------------------------------------------------------------------

class TestLoadRules:
    def test_returns_correct_count(self, simple_workbook):
        with ExcelReader(simple_workbook) as reader:
            rules = reader.load_rules()
        assert len(rules) == 2

    def test_rule_fields(self, simple_workbook):
        with ExcelReader(simple_workbook) as reader:
            rules = reader.load_rules()
        rule = next(r for r in rules if r.key == "cs-S1006")
        assert rule.name == "Method overrides should call super()"
        assert rule.severity == "CRITICAL"
        assert rule.language == "cs"
        assert rule.sheet_iden == "S1006"
        assert rule.count == 1


# ---------------------------------------------------------------------------
# Tests — load_issues_for_rule
# ---------------------------------------------------------------------------

class TestLoadIssuesForRule:
    def test_open_issues_only(self, simple_workbook):
        with ExcelReader(simple_workbook) as reader:
            rules = reader.load_rules()
            rule_s1110 = next(r for r in rules if r.key == "cs-S1110")
            issues = reader.load_issues_for_rule(rule_s1110)
        # uuid-003 has status=RESOLVED — must be excluded
        assert len(issues) == 1
        assert issues[0].key == "uuid-002"

    def test_issue_fields(self, simple_workbook):
        with ExcelReader(simple_workbook) as reader:
            rules = reader.load_rules()
            rule = next(r for r in rules if r.key == "cs-S1006")
            issues = reader.load_issues_for_rule(rule)
        assert len(issues) == 1
        issue = issues[0]
        assert issue.key == "uuid-001"
        assert issue.severity == "CRITICAL"
        assert issue.line == 42
        assert issue.component == "myrepo:main:src/Foo.cs"
        assert issue.status == "OPEN"
        assert issue.rule is rule


# ---------------------------------------------------------------------------
# Tests — load_all_issues
# ---------------------------------------------------------------------------

class TestLoadAllIssues:
    def test_total_open_issues(self, simple_workbook):
        with ExcelReader(simple_workbook) as reader:
            issues = reader.load_all_issues()
        # uuid-001 (OPEN) + uuid-002 (OPEN); uuid-003 is RESOLVED
        assert len(issues) == 2

    def test_severity_sorted(self, simple_workbook):
        with ExcelReader(simple_workbook) as reader:
            issues = reader.load_all_issues()
        # CRITICAL before MAJOR
        assert issues[0].severity == "CRITICAL"
        assert issues[1].severity == "MAJOR"

    def test_allowed_rules_filter(self, simple_workbook):
        with ExcelReader(simple_workbook) as reader:
            issues = reader.load_all_issues(allowed_rules={"cs-S1006"})
        assert all(i.rule.key == "cs-S1006" for i in issues)
        assert len(issues) == 1

    def test_severity_threshold(self, simple_workbook):
        with ExcelReader(simple_workbook, severity_threshold="CRITICAL") as reader:
            issues = reader.load_all_issues()
        # Only CRITICAL issues pass the threshold; MAJOR is skipped
        assert all(i.severity == "CRITICAL" for i in issues)


# ---------------------------------------------------------------------------
# Tests — error handling
# ---------------------------------------------------------------------------

class TestEdgeCases:
    def test_missing_sheet_does_not_crash(self, tmp_path):
        """If sheetIden points to a non-existent sheet, it is skipped gracefully."""
        buf = _make_workbook(
            rules=[("cs-S9999", "NONEXISTENT", "Ghost rule", "MINOR", "cs", "noaction", 0)],
            issues_per_rule={},  # no corresponding sheet
        )
        p = tmp_path / "ghost.xlsx"
        p.write_bytes(buf.read())
        with ExcelReader(p) as reader:
            rules = reader.load_rules()
            issues = reader.load_issues_for_rule(rules[0])
        assert issues == []

    def test_sheet_with_wrong_columns_is_skipped(self, tmp_path):
        """A sheet that exists but lacks required columns is skipped gracefully."""
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        inst = wb.active
        inst.title = "Instructions"

        rules_ws = wb.create_sheet("Rules")
        rules_ws.append(["key", "sheetIden", "name", "severity", "language", "action", "count"])
        rules_ws.append(["cs-S1234", "BadSheet", "Bad rule", "MAJOR", "cs", "noaction", 1])

        bad_ws = wb.create_sheet("BadSheet")
        # Missing the required 'key', 'severity', 'status' columns
        bad_ws.append(["foo", "bar", "baz"])
        bad_ws.append(["val1", "val2", "val3"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        p = tmp_path / "bad_columns.xlsx"
        p.write_bytes(buf.read())

        with ExcelReader(p) as reader:
            rules = reader.load_rules()
            issues = reader.load_issues_for_rule(rules[0])
        assert issues == []

    def test_load_all_issues_continues_past_bad_sheet(self, tmp_path):
        """load_all_issues skips bad sheets and still returns issues from good ones."""
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        inst = wb.active
        inst.title = "Instructions"

        rules_ws = wb.create_sheet("Rules")
        rules_ws.append(["key", "sheetIden", "name", "severity", "language", "action", "count"])
        rules_ws.append(["cs-S0001", "MISSING_SHEET", "Bad rule", "MAJOR", "cs", "noaction", 1])
        rules_ws.append(["cs-S0002", "GoodSheet", "Good rule", "MAJOR", "cs", "noaction", 1])

        good_ws = wb.create_sheet("GoodSheet")
        good_ws.append(["key", "severity", "message", "line", "component", "assigneeU", "assignee", "status"])
        good_ws.append(["uuid-ok", "MAJOR", "Fix me", 5, "repo:main:src/X.cs", "", "", "OPEN"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        p = tmp_path / "mixed.xlsx"
        p.write_bytes(buf.read())

        with ExcelReader(p) as reader:
            issues = reader.load_all_issues()

        assert len(issues) == 1
        assert issues[0].key == "uuid-ok"
